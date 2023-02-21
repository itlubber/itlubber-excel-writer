# -*- coding: utf-8 -*-
"""
@Time    : 2023/2/14 16:23
@Author  : itlubber
@Site    : itlubber.art
"""
import re
import os

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from openpyxl.drawing.image import Image
from openpyxl import load_workbook, Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import NamedStyle, Border, Side, Alignment, PatternFill, Font


class ExcelWriter:

    def __init__(self, style_excel='报告输出模版.xlsx', style_sheet_name="初始化", fontsize=10, font='楷体', theme_color='8E8BFE'):
        # english_width，chinese_width
        self.english_width = 0.12
        self.chinese_width = 0.21
        self.theme_color = theme_color
        self.fontsize = 10
        self.font = '楷体'

        self.workbook = load_workbook(style_excel)
        self.style_sheet = self.workbook[style_sheet_name]

        self.name_styles = []
        self.init_style(font, fontsize, theme_color)
        for style in self.name_styles:
            if style.name not in self.workbook.style_names:
                self.workbook.add_named_style(style)

    def add_conditional_formatting(self, worksheet, start_space, end_space):
        worksheet.conditional_formatting.add(f'{start_space}:{end_space}', DataBarRule(start_type='min', end_type='max', color=self.theme_color))

    @staticmethod
    def set_column_width(worksheet, column, width):
        worksheet.column_dimensions[column if isinstance(column, str) else get_column_letter(column)] = width

    def get_sheet_by_name(self, name):
        if name not in self.workbook.sheetnames:
            worksheet = self.workbook.copy_worksheet(self.style_sheet)
            worksheet.title = name
        else:
            worksheet = self.workbook[name]

        return worksheet

    def insert_value2sheet(self, worksheet, insert_space, value="", style="content", auto_width=False):
        if isinstance(insert_space, str):
            worksheet[insert_space] = value
            cell = worksheet[insert_space]
            start_col = re.findall('\D+', insert_space)[0]
            start_row = int(re.findall("\d+", insert_space)[0])
        else:
            cell = worksheet.cell(insert_space[0], insert_space[1], value)
            start_col = get_column_letter(insert_space[1])
            start_row = insert_space[0]
        cell.style =  style

        if auto_width:
            curr_width = worksheet.column_dimensions[start_col].width
            auto_width = min(max([(self.check_contain_chinese(value)[1] * self.english_width + self.check_contain_chinese(value)[2] * self.chinese_width) * self.fontsize, 10, curr_width]), 50)
            worksheet.column_dimensions[start_col].width = auto_width

        return start_row + 1, column_index_from_string(start_col) + 1

    def insert_pic2sheet(self, worksheet, fig, insert_space, figsize=(600, 250)):
        if isinstance(insert_space, str):
            start_row = int(re.findall("\d+", insert_space)[0])
            start_col = re.findall('\D+', insert_space)[0]
        else:
            start_row, start_col = insert_space
            start_col = get_column_letter(start_col)

        image = Image(fig)
        image.width, image.height = figsize
        worksheet.add_image(image, f"{start_col}{start_row}")

        return start_row + int(figsize[1] / 17.5), column_index_from_string(start_col) + 8

    def insert_rows(self, worksheet, row, row_index, col_index, merge_rows=None, style="", auto_width=False):
        curr_col = column_index_from_string(col_index)
        for j, v in enumerate(row):
            if merge_rows is not None and row_index + 1 not in merge_rows:
                if j == 0:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style="merge_left", auto_width=auto_width)
                elif j == len(row) - 1:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style="merge_right", auto_width=auto_width)
                else:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style="merge_middle", auto_width=auto_width)
            else:
                if j == 0:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=f"{style}_left" if style else "left", auto_width=auto_width)
                elif j == len(row) - 1:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=f"{style}_right" if style else "right", auto_width=auto_width)
                else:
                    self.insert_value2sheet(worksheet, f'{get_column_letter(curr_col + j)}{row_index}', self.astype_insertvalue(v), style=f"{style}_middle" if style else "middle", auto_width=auto_width)

    def insert_df2sheet(self, worksheet, data, insert_space, merge_column=None, header=True, index=False, auto_width=False):
        df = data.copy()

        if isinstance(insert_space, str):
            start_row = int(re.findall("\d+", insert_space)[0])
            start_col = re.findall('\D+', insert_space)[0]
        else:
            start_row, start_col = insert_space
            start_col = get_column_letter(start_col)

        if merge_column:
            if isinstance(merge_column, str):
                merge_column = [merge_column]

            if isinstance(merge_column[0], (int, float)):
                merge_cols = None
                merge_rows = merge_rows
            else:
                merge_cols = [get_column_letter(df.columns.get_loc(col) + column_index_from_string(start_col)) for col in merge_column]
                df = df.sort_values(merge_column)
                merge_rows = list(np.cumsum(df.groupby(merge_column)[merge_column].count().values[:, 0]) + start_row + 1)

        for i, row in enumerate(dataframe_to_rows(df, header=header, index=index)):
            if i == 0:
                if header:
                    self.insert_rows(worksheet, row, start_row + i, start_col, style="header", auto_width=auto_width)
                else:
                    self.insert_rows(worksheet, row, start_row + i, start_col, style="first", auto_width=auto_width)
            elif (header and i == len(df)) or (not header and i + 1 == len(df)):
                self.insert_rows(worksheet, row, start_row + i, start_col, style="last", auto_width=auto_width)
            else:
                self.insert_rows(worksheet, row, start_row + i, start_col, auto_width=auto_width, merge_rows=merge_rows if merge_column else None)

        # if merge_column and merge_cols is not None:
        #     merge_rows = [start_row + 2] + merge_rows
        #     for s, e in zip(merge_rows[:-1], merge_rows[1:]):
        #         if e - s > 1:
        #             for merge_col in merge_cols:
        #                 worksheet.merge_cells(f"{merge_col}{s-1}:{merge_col}{e-1}")

        end_row = start_row + len(data) + 1 if header else start_row + len(data)

        return (end_row, column_index_from_string(start_col) + len(data.columns))

    @staticmethod
    def check_contain_chinese(check_str):
        out = []
        for ch in str(check_str).encode('utf-8').decode('utf-8'):
            if u'\u4e00' <= ch <= u'\u9fff':
                out.append(True)
            else:
                out.append(False)
        return out, len(out) - sum(out), sum(out)

    @staticmethod
    def astype_insertvalue(value, decimal_point=4):
        if re.search('tuple|list|numpy.dtype|bool|str|numpy.ndarray|Interval|Categorical', str(type(value))):
            value = str(value)
        elif re.search('int', str(type(value))):
            value = value
        elif re.search('float', str(type(value))):
            value = round(float(value), decimal_point)
        else:
            value = 'nan'

        return value

    @staticmethod
    def calc_continuous_cnt(list_, index_=0):
        """
        Clac continuous_cnt

        Examples:s
            list_ = ['A','A','A','A','B','C','C','D','D','D']
            (1) calc_continuous_cnt(list_, 0) ===>('A', 0, 4)
            (2) calc_continuous_cnt(list_, 4) ===>('B', 4, 1)
            (3) calc_continuous_cnt(list_, 6) ===>('C', 6, 1)
        """
        if index_ >= len(list_):
            return None, None, None

        else:
            cnt, str_ = 0, list_[index_]
            for i in range(index_, len(list_), 1):
                if list_[i] == str_:
                    cnt = cnt + 1
                else:
                    break
            return str_, index_, cnt

    @staticmethod
    def itlubber_border(border, color):
        if len(border) == 3:
            return Border(left=Side(border_style=border[0], color=color[0]), right=Side(border_style=border[1], color=color[1]), bottom=Side(border_style=border[2], color=color[2]),)
        else:
            return Border(left=Side(border_style=border[0], color=color[0]), right=Side(border_style=border[1], color=color[1]), bottom=Side(border_style=border[2], color=color[2]), top=Side(border_style=border[3], color=color[3]),)

    @staticmethod
    def get_cell_space(space):
        if isinstance(space, str):
            start_row = int(re.findall("\d+", space)[0])
            start_col = re.findall('\D+', space)[0]
            return start_row, column_index_from_string(start_col)
        else:
            start_row = space[0]
            if isinstance(space[1], int):
                start_col = get_column_letter(space[1])
            else:
                start_col = space[1]
            return f"{start_row}{start_col}"

    def init_style(self, font, fontsize, theme_color):
        header_style, header_left_style, header_middle_style, header_right_style = NamedStyle(name="header"), NamedStyle(name="header_left"), NamedStyle(name="header_middle"), NamedStyle(name="header_right")
        last_style, last_left_style, last_middle_style, last_right_style = NamedStyle(name="last"), NamedStyle(name="last_left"), NamedStyle(name="last_middle"), NamedStyle(name="last_right")
        content_style, left_style, middle_style, right_style = NamedStyle(name="content"), NamedStyle(name="left"), NamedStyle(name="middle"), NamedStyle(name="right")
        merge_style, merge_left_style, merge_middle_style, merge_right_style =  NamedStyle(name="merge"), NamedStyle(name="merge_left"), NamedStyle(name="merge_middle"), NamedStyle(name="merge_right")
        first_style, first_left_style, first_middle_style, first_right_style = NamedStyle(name="first"), NamedStyle(name="first_left"), NamedStyle(name="first_middle"), NamedStyle(name="first_right")

        header_font = Font(size=fontsize, name=font, color="FFFFFF", bold=True)
        header_fill = PatternFill(fill_type="solid", start_color=theme_color)
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
        content_fill = PatternFill(fill_type="solid", start_color="FFFFFF")
        content_font = Font(size=fontsize, name=font, color="000000")

        header_style.font, header_left_style.font, header_middle_style.font, header_right_style.font = header_font, header_font, header_font, header_font
        header_style.fill, header_left_style.fill, header_middle_style.fill, header_right_style.fill = header_fill, header_fill, header_fill, header_fill
        header_style.alignment, header_left_style.alignment, header_middle_style.alignment, header_right_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True), alignment, alignment, alignment

        header_style.border = self.itlubber_border(["medium", "medium", "medium", "medium"], [theme_color, theme_color, theme_color, theme_color])
        header_left_style.border = self.itlubber_border(["medium", "thin", "medium", "medium"], [theme_color, "FFFFFF", theme_color, theme_color])
        header_middle_style.border = self.itlubber_border(["thin", "thin", "medium", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color])
        header_right_style.border = self.itlubber_border(["thin", "medium", "medium", "medium"], ["FFFFFF", theme_color, theme_color, theme_color])

        last_style.font, last_left_style.font, last_middle_style.font, last_right_style.font = content_font, content_font, content_font, content_font
        last_style.fill, last_left_style.fill, last_middle_style.fill, last_right_style.fill = content_fill, content_fill, content_fill, content_fill
        last_style.alignment, last_left_style.alignment, last_middle_style.alignment, last_right_style.alignment = alignment, alignment, alignment, alignment

        last_style.border = self.itlubber_border(["medium", "medium", "medium"], [theme_color, theme_color, theme_color])
        last_left_style.border = self.itlubber_border(["medium", "thin", "medium"], [theme_color, "FFFFFF", theme_color])
        last_middle_style.border = self.itlubber_border(["thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color])
        last_right_style.border = self.itlubber_border(["thin", "medium", "medium"], ["FFFFFF", theme_color, theme_color])

        content_style.font, left_style.font, middle_style.font, right_style.font = content_font, content_font, content_font, content_font
        content_style.fill, left_style.fill, middle_style.fill, right_style.fill = content_fill, content_fill, content_fill, content_fill
        content_style.alignment, left_style.alignment, middle_style.alignment, right_style.alignment = alignment, alignment, alignment, alignment

        content_style.border = self.itlubber_border(["medium", "medium", "thin"], [theme_color, theme_color, theme_color])
        left_style.border = self.itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", theme_color])
        middle_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", "FFFFFF", theme_color])
        right_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, theme_color])

        merge_style.font, merge_left_style.font, merge_middle_style.font, merge_right_style.font = content_font, content_font, content_font, content_font
        merge_style.fill, merge_left_style.fill, merge_middle_style.fill, merge_right_style.fill = content_fill, content_fill, content_fill, content_fill
        merge_style.alignment, merge_left_style.alignment, merge_middle_style.alignment, merge_right_style.alignment = alignment, alignment, alignment, alignment

        merge_style.border = self.itlubber_border(["medium", "medium", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        merge_left_style.border = self.itlubber_border(["medium", "thin", "thin"], [theme_color, "FFFFFF", "FFFFFF"])
        merge_middle_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", "FFFFFF", "FFFFFF"])
        merge_right_style.border = self.itlubber_border(["thin", "medium", "thin"], ["FFFFFF", theme_color, "FFFFFF"])

        first_style.font, first_left_style.font, first_middle_style.font, first_right_style.font = content_font, content_font, content_font, content_font
        first_style.fill, first_left_style.fill, first_middle_style.fill, first_right_style.fill = content_fill, content_fill, content_fill, content_fill
        first_style.alignment, first_left_style.alignment, first_middle_style.alignment, first_right_style.alignment = alignment, alignment, alignment, alignment

        first_style.border = self.itlubber_border(["medium", "medium", "thin", "medium"], [theme_color, theme_color, theme_color, theme_color])
        first_left_style.border = self.itlubber_border(["medium", "thin", "thin", "medium"], [theme_color, "FFFFFF", theme_color, theme_color])
        first_middle_style.border = self.itlubber_border(["thin", "thin", "thin", "medium"], ["FFFFFF", "FFFFFF", theme_color, theme_color])
        first_right_style.border = self.itlubber_border(["thin", "medium", "thin", "medium"], ["FFFFFF", theme_color, theme_color, theme_color])

        self.name_styles.extend([
            header_style, header_left_style, header_middle_style, header_right_style,
            last_style, last_left_style, last_middle_style, last_right_style,
            content_style, left_style, middle_style, right_style,
            merge_style, merge_left_style, merge_middle_style, merge_right_style,
            first_style, first_left_style, first_middle_style, first_right_style
        ])

    def save(self, filename):
        self.workbook.remove(self.style_sheet)
        self.workbook.save(filename)
        self.workbook.close()


if __name__ == '__main__':
    writer = ExcelWriter(style_excel="/Users/lubberit/Desktop/金融/兴业银行贷中行为评分/utils/报告输出模版.xlsx")
    worksheet = writer.get_sheet_by_name("模型报告")
    writer.insert_value2sheet(worksheet, "B2", value="模型报告", style="header")
    writer.insert_value2sheet(worksheet, "B3", value="当前模型主要为评分卡模型", style="content", auto_width=True)
    end_row = writer.insert_pic2sheet(worksheet, "/Users/lubberit/Desktop/金融/兴业银行贷中行为评分/tests/mypic.png", "B5")
    end_row = writer.insert_pic2sheet(worksheet, "/Users/lubberit/Desktop/金融/兴业银行贷中行为评分/tests/mypic.png", "H5")
    sample = pd.DataFrame(np.concatenate([np.random.random_sample((10, 10)) * 40, np.random.randint(0, 3, (10, 2))], axis=1), columns=[f"B{i}" for i in range(10)] + ["target", "type"])
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")))
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), merge_column="target")
    end_row, end_col = writer.insert_df2sheet(worksheet, sample, (end_row + 2, column_index_from_string("B")), merge_column=["target", "type"])
    writer.save("test.xlsx")
